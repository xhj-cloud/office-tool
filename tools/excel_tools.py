"""
Excel 电子表格读写工具
支持 .xlsx / .xlsm 格式的读取和生成
"""

import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from .json_repair import safe_parse_json


def _get_cell_value(cell):
    """安全获取单元格值"""
    if cell.value is None:
        return ""
    if isinstance(cell.value, float):
        # 处理日期序列号
        return cell.value
    return str(cell.value)


def read_xlsx(file_path: str, sheet_name: str = None, mode: str = "full") -> str:
    """
    读取 Excel 文件内容

    Args:
        file_path: .xlsx 文件路径
        sheet_name: 工作表名称（默认读取第一个 sheet）
        mode: "full"(全部) / "structure"(含样式) / "values"(仅数据)

    Returns:
        JSON 格式的内容
    """
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        return json.dumps({"error": f"无法打开文件: {str(e)}"}, ensure_ascii=False)

    # 确定 sheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return json.dumps({"error": f"Sheet '{sheet_name}' 不存在，可用: {wb.sheetnames}"}, ensure_ascii=False)
        ws = wb[sheet_name]
    else:
        ws = wb.active

    output = {
        "file": file_path,
        "sheet": ws.title,
        "sheets_available": wb.sheetnames,
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
    }

    if mode == "values":
        # 仅数据
        data = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
            data.append([str(v) if v is not None else "" for v in row])
        output["data"] = data

    elif mode == "structure":
        # 含样式
        data = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            row_data = []
            for cell in row:
                cell_info = {
                    "value": _get_cell_value(cell),
                    "column": cell.column_letter,
                    "row": cell.row,
                    "font": {
                        "name": cell.font.name,
                        "size": cell.font.size,
                        "bold": cell.font.bold,
                    } if cell.font else None,
                    "fill": cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None,
                }
                row_data.append(cell_info)
            data.append(row_data)
        output["data"] = data

    else:  # full
        data = []
        merged_cells = [str(m) for m in ws.merged_cells.ranges]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            data.append([_get_cell_value(cell) for cell in row])
        output["data"] = data
        output["merged_cells"] = merged_cells

    # 检测无缓存值的公式单元格：data_only=True 会把它们读成空，必须明确告知
    try:
        wb_f = load_workbook(file_path, data_only=False)
        ws_f = wb_f[ws.title] if ws.title in wb_f.sheetnames else None
        if ws_f is not None:
            uncached = []
            for row in ws_f.iter_rows(min_row=1, max_row=ws_f.max_row, max_col=ws_f.max_column):
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        if ws.cell(row=cell.row, column=cell.column).value is None:
                            uncached.append(f"{cell.coordinate}: {v}")
            wb_f.close()
            if uncached:
                output["uncached_formula_cells"] = {
                    "note": "以下公式单元格没有缓存值（文件从未在 Excel 中打开计算过），因此读出来是空的",
                    "cells": uncached,
                }
    except Exception:
        pass  # 检测失败不影响主流程

    wb.close()
    return json.dumps(output, ensure_ascii=False, indent=2)


def write_xlsx(spec_json: str) -> str:
    """
    根据 JSON 规格生成 .xlsx 文件

    spec_json 格式:
    {
        "output": "/path/to/output.xlsx",
        "sheets": [
            {
                "name": "Sheet1",
                "headers": ["序号", "名称", "金额"],   // 表头（自动加粗、居中、蓝底白字）
                "rows": [["1", "交换机", 3995]],       // 数据行
                "col_widths": [8, 30, 15],             // 列宽（字符数）
                "merged_cells": ["A1:B1"],             // 合并单元格
                "freeze": "A2"                          // 冻结窗格
            }
        ]
    }
    """
    spec, err = safe_parse_json(spec_json)
    if err:
        return json.dumps({"error": f"JSON 解析失败: {err}"}, ensure_ascii=False)

    output_path = spec.get("output")
    if not output_path:
        return json.dumps({"error": "必须指定 output 路径"}, ensure_ascii=False)

    wb = Workbook()

    # 默认样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="宋体", size=10.5)
    body_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for i, sheet_spec in enumerate(spec.get("sheets", [])):
        name = sheet_spec.get("name") or f"Sheet{i + 1}"
        if i == 0:
            # 复用默认空 sheet，避免残留多余的 "Sheet"
            ws = wb.active
            ws.title = name
        else:
            # Excel 不允许同名工作表，重名时自动追加序号（而不是误删用户 sheet）
            base, n = name, 2
            while name in wb.sheetnames:
                name = f"{base}{n}"
                n += 1
            ws = wb.create_sheet(title=name)

        headers = sheet_spec.get("headers", [])
        rows = sheet_spec.get("rows", [])
        col_widths = sheet_spec.get("col_widths", [])

        # 写表头
        if headers:
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

        # 写数据
        start_row = 2 if headers else 1
        for ri, row_data in enumerate(rows):
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=start_row + ri, column=ci, value=val)
                cell.font = body_font
                cell.alignment = center_align if ci == 1 else body_align
                cell.border = thin_border

        # 列宽
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # 合并单元格
        for mc in sheet_spec.get("merged_cells", []):
            ws.merge_cells(mc)

        # 冻结窗格
        if sheet_spec.get("freeze"):
            ws.freeze_panes = sheet_spec["freeze"]

    wb.save(output_path)
    return json.dumps({"success": True, "output": output_path, "sheets": wb.sheetnames}, ensure_ascii=False)


def edit_xlsx(file_path: str, spec_json: str) -> str:
    """
    编辑已有 Excel 文件（.xlsx / .xlsm）

    spec_json 格式:
    {
        "output": "/path/to/output.xlsx",   // 可选，不指定则覆盖原文件
        "sheet": "Sheet1",                   // 可选，默认第一个 sheet
        "operations": [
            {"op": "set_cells", "values": {"A1": "新值", "B2": 123}},      // 批量写单元格值
            {"op": "set_cell", "cell": "A1", "value": "...", "bold": true, "font_size": 12},
            {"op": "append_rows", "rows": [["a","b"],["c","d"]]},           // 追加多行
            {"op": "delete_rows", "start": 2, "count": 3},                  // 删除行
            {"op": "merge_cells", "range": "A1:C1"},                        // 合并单元格
            {"op": "unmerge_cells", "range": "A1:C1"},                      // 取消合并
            {"op": "rename_sheet", "new_name": "新名称"}                     // 重命名当前 sheet
        ]
    }
    """
    spec, err = safe_parse_json(spec_json)
    if err:
        return json.dumps({"error": f"JSON 解析失败: {err}"}, ensure_ascii=False)

    # .xlsm/.xlam 含 VBA 宏，必须 keep_vba 加载，否则保存时全部宏代码丢失
    keep_vba = str(file_path).lower().endswith((".xlsm", ".xlam"))
    try:
        wb = load_workbook(file_path, keep_vba=keep_vba)
    except Exception as e:
        return json.dumps({"error": f"无法打开文件: {str(e)}"}, ensure_ascii=False)

    ws_name = spec.get("sheet")
    if ws_name:
        if ws_name not in wb.sheetnames:
            wb.close()
            return json.dumps({"error": f"Sheet '{ws_name}' 不存在，可用: {wb.sheetnames}"}, ensure_ascii=False)
        ws = wb[ws_name]
    else:
        ws = wb.active

    for op in spec.get("operations", []):
        kind = op.get("op")
        try:
            if kind == "set_cells":
                for cell_ref, val in op.get("values", {}).items():
                    ws[cell_ref] = val

            elif kind == "set_cell":
                cell_ref = op.get("cell")
                ws[cell_ref] = op.get("value")
                bold = op.get("bold")
                fsize = op.get("font_size")
                if bold is not None or fsize:
                    # 合并现有字体属性，只覆盖指定的项（避免整体重置丢失字体名/字号等）
                    old = ws[cell_ref].font
                    ws[cell_ref].font = Font(
                        name=old.name,
                        size=fsize if fsize else old.size,
                        bold=bold if bold is not None else old.bold,
                        italic=old.italic,
                        underline=old.underline,
                        strike=old.strike,
                        color=old.color,
                    )

            elif kind == "append_rows":
                for row_data in op.get("rows", []):
                    ws.append(row_data)

            elif kind == "delete_rows":
                ws.delete_rows(op.get("start", 1), op.get("count", 1))

            elif kind == "merge_cells":
                ws.merge_cells(op.get("range"))

            elif kind == "unmerge_cells":
                ws.unmerge_cells(op.get("range"))

            elif kind == "rename_sheet":
                new = op.get("new_name")
                if new and new not in wb.sheetnames:
                    ws.title = new

            else:
                wb.close()
                return json.dumps({"error": f"未知操作: {kind}"}, ensure_ascii=False)

        except Exception as e:
            wb.close()
            return json.dumps({"error": f"操作 '{kind}' 执行失败: {str(e)}"}, ensure_ascii=False)

    output_path = spec.get("output", file_path)
    if keep_vba and not str(output_path).lower().endswith((".xlsm", ".xlam")):
        wb.close()
        return json.dumps({
            "error": f"源文件含 VBA 宏，不能保存为 {os.path.splitext(str(output_path))[1] or output_path}（会丢失全部宏代码），请指定 .xlsm/.xlam 输出路径",
        }, ensure_ascii=False)
    wb.save(output_path)
    wb.close()
    return json.dumps({"success": True, "output": output_path}, ensure_ascii=False)
